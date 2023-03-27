import openpyxl
import requests
from bs4 import BeautifulSoup
import re

# Prompt the user to enter the names of the input and output files
input_file = input("Enter the name of the input file: ")
output_file = input("Enter the name of the output file: ")

# Load the input Excel file
try:
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
except Exception as e:
    print(f"Error loading input file: {e}")
    exit()

# Create the output Excel file
try:
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = "College Info"
    output_ws.append(["College Name", "Website", "Address", "Principal Name"])
except Exception as e:
    print(f"Error creating output file: {e}")
    exit()

# Loop through each row in the input Excel file
for row in ws.iter_rows(min_row=2, values_only=True):
    college_name = row[0]
    url = row[1]

    # Make a GET request to the URL and parse the HTML using BeautifulSoup
    try:
        response = requests.get(url)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, "html.parser")
    except Exception as e:
        print(f"Error fetching {url}: {e}")
        continue

    # Find the address element on the page using a regular expression
    regex = re.compile(r"address|contact", re.IGNORECASE)
    address_elem = soup.find_all("div", text=regex)

    if len(address_elem) == 0:
        print(f"No address found on {url}")
        continue

    # Extracting the text of the address element
    address = address_elem[0].get_text().strip()

    # Finding the principal name element on the page using a regular expression
    regex = re.compile(r"principal", re.IGNORECASE)
    principal_elem = soup.find_all(text=regex)

    if len(principal_elem) == 0:
        principal_name = ""
    else:
        # Extract the text of the principal name element
        principal_name = principal_elem[0].strip()

   
    output_ws.append([college_name, url, address, principal_name])

# Save the output Excel file
try:
    output_wb.save(output_file)
except Exception as e:
    print(f"Error saving output file: {e}")
    exit()
